import openpyxl as oxl
import pulp as plp
import timeit
import pandas as pd

start = timeit.default_timer()

###### PROBLEM PARAMETERS

##### THE DEFAULT CONFIGURATION WILL FIND YOU THE SMALLEST-SIZED COMBINATION OF DRIVERS THAT GRANTS FULL TOP SHELF COVERAGE, FOR A PERSONALIZED OUTPUT BASED ON YOUR INVENTORY
##### PLEASE FOLLOW THE README'S PARAMETERS GUIDELINES

TARGET_LEVEL_7 = False
HIGH_ENDS_ONLY = False
ONLY_USE_ADQUIRED_DRIVERS = False
MINIMUM_DRIVERS_PER_COURSE = 2
CONSIDER_CURRENT_INVESTMENT_LEVELS = False
REQUIRE_PRIORITY_DRIVERS = False
MINIMUM_PRIORITY_DRIVERS_PER_COURSE = 2 
UNAQUIRED_DRIVERS_COST = 2 # Measured in tickets
PRIORITY_DRIVERS = ['Bowser (Santa)', 'Dry Bones (Gold)', 
'Gold Koopa (Freerunning)', 'King Bob-omb (Gold)', 'King Boo (Gold)',
'Mario (Hakama)', 'Mario (Tuxedo)', 'Pauline (Party Time)'
'Peach (Vacation)', 'Pink Gold Peach', 'Rosalina (Swimwear)',
'Shy Guy (Ninja)', 'Mario (Sunshine)', 
'Boomerang Bro', 'Donkey Kong', 'Toad (Pit Crew)']
FIND_ALL_COMBINATIONS = Fale

###### LOAD DATA

# Tickets and level
cost = {
'N':{'cost' : 800,
     'requirements' : {0:40 + UNAQUIRED_DRIVERS_COST, 1:40, 2:38, 3:33, 4:25, 5:14, 6:0, 7:20}
     },
'S':{'cost' : 3000,
     'requirements' : {0:15 + UNAQUIRED_DRIVERS_COST, 1:15, 2:14, 3:12, 4:9, 5:5, 6:0, 7:8}
     },
'H':{'cost' : 12000,
     'requirements' : {0:9 + UNAQUIRED_DRIVERS_COST, 1:9, 2:8, 3:7, 4:5, 5:3, 6:0, 7:5}
     }
}

wb = oxl.load_workbook('raw.xlsx')
ws = wb['base']

courses = set()
drivers = set()
coverage = {}
investment = {}
costs = {}

# Function that calculates a drivers cost to max out

def calculate_driver_cost(tier, driver, current_level, current_tickets):
    
    if( HIGH_ENDS_ONLY and (tier == 'H' or driver in PRIORITY_DRIVERS) ) or not HIGH_ENDS_ONLY:
        calculated_cost = -1
        
        if current_level < 6:    
            calculated_cost +=  cost[tier]['cost'] * ( cost[tier]['requirements'][current_level] - current_tickets )
        
        calculated_cost +=  cost[tier]['cost'] * (cost[tier]['requirements'][7]) if TARGET_LEVEL_7 and current_level < 7 else 0
    
        if ( ONLY_USE_ADQUIRED_DRIVERS and current_level >= 1 ) or not ONLY_USE_ADQUIRED_DRIVERS :
            drivers.add(driver)
            costs[driver] = calculated_cost if CONSIDER_CURRENT_INVESTMENT_LEVELS else 1
        
    
# Load driver and iventory data    
        
row = 1
while( ws.cell(row=row, column=5).value is not None ):
    
    tier = ws.cell(row=row, column=4).value
    driver = ws.cell(row=row, column=5).value
    current_level = ws.cell(row=row, column=6).value
    current_tickets = ws.cell(row=row, column=7).value
    
    calculate_driver_cost(tier, driver, current_level, current_tickets)

    row += 1

# Read the course-driver coverage data

row = 1
while( ws.cell(row=row, column=1).value is not None ):
    
    driver = ws.cell(row=row, column=1).value.split(':')[0]
    course = ws.cell(row=row, column=2).value
    
    if str(course)[0] not in ('0', '^') and driver in drivers:
        
        if course in courses:
            coverage[course].add(driver)
        else:
            courses.add(course)
            coverage[course] = set([driver])
    
    row+= 1
    
# Remove unused priority drivers
prio_copy = set(PRIORITY_DRIVERS)
for d in prio_copy:
    if d not in drivers:
        PRIORITY_DRIVERS.remove(d)

# Minimum Drivers per Course
courses_copy = set(courses)
priority_courses = set()
for course in courses_copy:

    # Remove courses without the minimum required drivers
    if len(coverage[course]) < MINIMUM_DRIVERS_PER_COURSE:
        courses.remove(course)
        
    # Calculate if a course has enough priority drivers
    n = sum([ 1 if d in PRIORITY_DRIVERS else 0 for d in coverage[course] ])
    if n >= MINIMUM_PRIORITY_DRIVERS_PER_COURSE:
        priority_courses.add(course)


# Create the LP problem
model = plp.LpProblem('MKT_Maximum_Coverage', plp.LpMinimize)

chosen_drivers = {d : plp.LpVariable(f'{d}', lowBound=0, upBound=1, cat='Integer') for d in drivers}

# objective function
model += plp.lpSum( costs[d] * chosen_drivers[d] for d in drivers )

# Constraint: Cover each course at least once
for c in courses:
    if not REQUIRE_PRIORITY_DRIVERS or c not in priority_courses:
        model += plp.lpSum( chosen_drivers[d] for d in coverage[c] ) >= 1
    elif REQUIRE_PRIORITY_DRIVERS and c in priority_courses:
        model += plp.lpSum( chosen_drivers[d] for d in PRIORITY_DRIVERS if d in coverage[c] ) >= 1

#model+= chosen_drivers['Pauline (Party Time)'] == 0
#model+= chosen_drivers['Dry Bowser'] == 1

model.solve()
objVal = plp.value(model.objective)
iteration = 1
while objVal == plp.value(model.objective):
    
    # Solve the LP model
    print( plp.LpStatus[model.status] )
    print( plp.value(model.objective))
    
    # Print the solution
    OPTIMAL_DRIVERS = set()
    for d in drivers:
        if chosen_drivers[d].varValue == 1:
            OPTIMAL_DRIVERS.add(d)
            
    output = []
    for d in OPTIMAL_DRIVERS:
        
        unique_courses = set()
        for c in courses:
            if d in coverage[c]:
                n = sum([ 1 if d2 in coverage[c] else 0 for d2 in OPTIMAL_DRIVERS ])
                if n == 1:
                    unique_courses.add(c)
                if d in PRIORITY_DRIVERS:
                    n = sum([ 1 if d2 in coverage[c] else 0 for d2 in OPTIMAL_DRIVERS.intersection(PRIORITY_DRIVERS) ])
                    if n == 1:
                        unique_courses.add(c)
        x = 1
        if len(unique_courses) > 0:
            row = {
                   'combination' : iteration,
                   'driver' : d, 'cost':costs[d]+1, 
                   'unique_coverage': len(unique_courses),
                   'value': (costs[d]+1)/len(unique_courses),
                   'unique_courses': str(unique_courses)
                   }
            output.append(row)
        #print(f'{d} costs {costs[d]} to max out and covers {len(unique_courses)} unique courses: \n{unique_courses}')
        if FIND_ALL_COMBINATIONS:
            model += plp.lpSum( chosen_drivers[d] for d in OPTIMAL_DRIVERS) <= len(OPTIMAL_DRIVERS) - 1
            model.solve()
        else:
            objVal = -1
        
        ans_df = pd.DataFrame(output).sort_values(by=['value'])   
        ans_df.to_csv('drivers_unique.csv')
        
        iteration += 1
            

stop = timeit.default_timer()

print('Time: ', stop - start)  
