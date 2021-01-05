import openpyxl as oxl
from gurobipy import *

# Constants
DRIVERS = 105
COURSES = 148

# open the data file
wb = oxl.load_workbook('input.xlsx')
ws = wb['raw_matrix']

m = Model()
m.setParam('OutputFlag', 0)

# Drivers
D = [ ws.cell(row=1, column=2+d).value for d in range(DRIVERS)]
x = { d : m.addVar(vtype=GRB.BINARY) for d in D }

# Courses
C = [ ws.cell(row=2+c, column=1).value for c in range(COURSES)]

# Subsets
S = {}
for i in range(COURSES):
    s = []
    for j in range(DRIVERS):
        if ws.cell(row=2+i, column=2+j).value == 1:
            s.append(D[j])
    S[C[i]] = s

# Set constraints
for course in C:
    m.addConstr( quicksum( x[d] for d in S[course] ) >= 1)

m.setObjective( quicksum(x[d] for d in D), GRB.MINIMIZE)
m.optimize()

opt = m.objVal
best = m.objVal

number = 1
while opt == best:

    # Print solution
    print(f'With these {opt} drivers: ({number})')
    u = []
    row = 1
    for d in D:
        if x[d].x == 1:
            wb['solutions'].cell(row=row, column=number).value = d
            u.append(d)
            print(d)
            row += 1

    number += 1
    opt = m.objVal

    # Block solution
    m.addConstr(quicksum(x[d] for d in u) <= len(u) - 1)

    # Find new solution
    m.update()
    m.optimize()

wb.save('solution.xlsx')